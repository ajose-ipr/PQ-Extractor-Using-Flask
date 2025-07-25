import re
import pdfplumber
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# Set up logging
logger = logging.getLogger(__name__)

# Constants
EXPECTED_HARMONICS = set(range(2, 51))
FAIL_COLOR = 'color: darkred; font-weight: bold; background-color: #ffebeb'
HIGHLIGHT_COLOR = 'font-weight: bold; background-color: #fff3cd'

# Column definitions for all table types
VOLTAGE_COLUMNS = [
    "Harmonic", "Time Percent Limit[%]", "Reg Max[%]",
    "Measured_V1N", "Measured_V2N", "Measured_V3N", 
    "Result_V1N", "Result_V2N", "Result_V3N", "Page_Number"
]

CURRENT_COLUMNS = [
    "Harmonic", "Time Percent Limit[%]", "Reg Max[%]",
    "Measured_I1", "Measured_I2", "Measured_I3", 
    "Result_I1", "Result_I2", "Result_I3", "Page_Number"
]

# Updated section boundaries for all tables
SECTION_BOUNDARIES = {
    "HARMONIC VOLTAGE FULL TIME RANGE": [
        "SUMMARY", "TOTAL HARMONIC VOLTAGE FULL TIME RANGE", 
        "TOTAL HARMONIC DISTORTION FULL TIME RANGE", "HARMONIC CURRENT FULL TIME RANGE"
    ],
    "HARMONIC CURRENT FULL TIME RANGE": [
        "TOTAL HARMONIC DISTORTION DAILY", "TDD FULL TIME RANGE",
        "HARMONIC VOLTAGE DAILY", "TRANSIENT"
    ],
    "HARMONIC VOLTAGE DAILY": [
        "TOTAL HARMONIC DISTORTION FULL TIME RANGE", 
        "TOTAL HARMONIC VOLTAGE FULL TIME RANGE", "HARMONIC CURRENT DAILY", "TOTAL HARMONIC DISTORTION DAILY"
    ],
    "HARMONIC CURRENT DAILY": [
        "TDD FULL TIME RANGE", "TDD DAILY", "TRANSIENT", "FLICKER SEVERITY"
    ]
}

# All supported table names
SUPPORTED_TABLES = [
    "Harmonic Voltage Full Time Range",
    "Harmonic Current Full Time Range", 
    "Harmonic Voltage Daily",
    "Harmonic Current Daily"
]

# Enhanced regex patterns for text extraction
TEXT_EXTRACTION_PATTERNS = [
    # Standard pattern with Pass/Fail
    re.compile(
        r'(\d+)\s*,?\s*'  # Harmonic
        r'(\d+)\s*,?\s*'  # Time percent
        r'([\d.]+)\s*,?\s*'  # Reg max
        r'([\d.]+)\s*,?\s*'  # Measured 1
        r'([\d.]+)\s*,?\s*'  # Measured 2
        r'([\d.]+)\s*,?\s*'  # Measured 3
        r'(Pass|Fail)\s*\(([\d.%]+)\)\s*,?\s*'  # Result 1
        r'(Pass|Fail)\s*\(([\d.%]+)\)\s*,?\s*'  # Result 2
        r'(Pass|Fail)\s*\(([\d.%]+)\)',  # Result 3
        re.IGNORECASE
    ),
    # Pattern without explicit Pass/Fail words
    re.compile(
        r'(\d+)\s*,?\s*'
        r'(\d+)\s*,?\s*'
        r'([\d.]+)\s*,?\s*'
        r'([\d.]+)\s*,?\s*'
        r'([\d.]+)\s*,?\s*'
        r'([\d.]+)\s*,?\s*'
        r'\(([\d.%]+)\)\s*,?\s*'
        r'\(([\d.%]+)\)\s*,?\s*'
        r'\(([\d.%]+)\)',
        re.IGNORECASE
    ),
    # Pattern for multiline data
    re.compile(
        r'(\d+)\s*,?\s*(\d+)\s*,?\s*([\d.]+)\s*,?\s*([\d.]+)\s*,?\s*([\d.]+)\s*,?\s*([\d.]+)',
        re.IGNORECASE | re.MULTILINE
    )
]

# Add this helper function to processing.py
def is_7_day_report(filename):
    """Check if filename indicates a 7-day report"""
    filename_upper = str(filename).upper()
    return "7" in filename_upper and "DAY" in filename_upper


def safe_float_convert(val):
    """Convert value to float, preserve zeros and empty strings"""
    if val is None or val == "":
        return ""
    try:
        return float(val)
    except:
        return str(val)

def extract_metadata(pdf_file, filename):
    """Extract metadata from PDF file"""
    name = filename if isinstance(filename, str) else filename.name
    component_info = re.findall(r"\((.*?)\)", str(name))
    component_text = component_info[0] if component_info else "Not found"

    report_info = {
        "start_time": "Not found", 
        "end_time": "Not found", 
        "gmt": "Not found", 
        "version": "Not found"
    }

    try:
        with pdfplumber.open(pdf_file if isinstance(pdf_file, str) else pdf_file) as pdf:
            text0 = pdf.pages[0].extract_text() or ""
        
        time_pattern = re.compile(
            r"Start time:\s*(\d{2}-\d{2}-\d{4}\s*\d{2}:\d{2}:\d{2}\s*[AP]M)\s*"
            r"End time:\s*(\d{2}-\d{2}-\d{4}\s*\d{2}:\d{2}:\d{2}\s*[AP]M)\s*"
            r"GMT:\s*([+-]\d{2}:\d{2})\s*"
            r"Report Version:\s*([\d.]+)"
        )
        
        match = time_pattern.search(text0)
        if match:
            report_info = {
                "start_time": match.group(1),
                "end_time": match.group(2),
                "gmt": match.group(3),
                "version": match.group(4)
            }
            
        combined_text = (str(name) + " " + text0).upper()
        block_match = re.search(r"\bBLOCK[-\s]*(\d{1,3})\b", combined_text)
        feeder_match = re.search(r"\b(FEEDER|BAY)[-\s]*(\d{1,3})\b", combined_text)
        company_match = re.search(r"\b(TATA|ADANI|NTPC|RELIANCE|POWERGRID|TORRENT)\b", combined_text)

        block = block_match.group(1) if block_match else "Not found"
        feeder = feeder_match.group(2) if feeder_match else "Not found"
        company = company_match.group(1) if company_match else "Not found"

        return component_text, block, feeder, company, report_info

    except Exception as e:
        logger.error(f"Error extracting metadata: {str(e)}")
        return "Not found", "Error", "Error", "Error", report_info

def extract_table_data_from_text(text, has_results=True, page_number=None):
    """Enhanced text extraction - preserve all values including zeros"""
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'(Pass|Fail)\s*\(\s*([\d.%]+)\s*\)', r'\1(\2)', text)
    
    data = []
    
    # More flexible patterns to catch all data
    flexible_patterns = [
        # Original patterns
        *TEXT_EXTRACTION_PATTERNS,
        # Additional flexible pattern for edge cases
        re.compile(
            r'(\d+)\s*[,\s]*(\d+)\s*[,\s]*([\d.]+)\s*[,\s]*([\d.\s]*)\s*[,\s]*([\d.\s]*)\s*[,\s]*([\d.\s]*)',
            re.IGNORECASE
        )
    ]
    
    for pattern in flexible_patterns:
        for match in pattern.finditer(text):
            try:
                harmonic = int(match.group(1))
                
                # FILTER: Skip harmonic 1 and invalid range
                if harmonic == 1 or harmonic < 2 or harmonic > 50 or harmonic > 1000:
                    continue
                
                groups = match.groups()
                
                if len(groups) >= 12:  # Full pattern with Pass/Fail
                    row = [
                        harmonic, match.group(2), match.group(3), 
                        match.group(4) or "", match.group(5) or "", match.group(6) or "",
                        f"{match.group(7)}({match.group(8)})",
                        f"{match.group(9)}({match.group(10)})", 
                        f"{match.group(11)}({match.group(12)})",
                        page_number
                    ]
                elif len(groups) >= 9:  # Pattern without Pass/Fail
                    row = [
                        harmonic, match.group(2), match.group(3),
                        match.group(4) or "", match.group(5) or "", match.group(6) or "",
                        f"Pass({match.group(7)})", f"Pass({match.group(8)})", 
                        f"Pass({match.group(9)})", page_number
                    ]
                elif len(groups) >= 6 and not has_results:  # Just measurements
                    row = [
                        harmonic, match.group(2), match.group(3),
                        match.group(4) or "", match.group(5) or "", match.group(6) or "",
                        "N/A", "N/A", "N/A", page_number
                    ]
                else:
                    continue
                    
                data.append(row)
            except (ValueError, IndexError):
                continue
    
    return data

def _extract_structured_data(page_tables, tables, active_table, page_number):
    """Helper function to extract structured table data with page tracking"""
    for table in page_tables:
        if len(table) > 1:
            for row in table:
                if row and str(row[0]).strip().isdigit():
                    try:
                        harmonic = int(row[0])
                        
                        # FILTER 1: Skip harmonic 1 (fundamental frequency)
                        if harmonic == 1:
                            continue
                            
                        # FILTER 2: Only accept valid harmonic range (2-50)
                        if harmonic < 2 or harmonic > 50:
                            continue
                            
                        # FILTER 3: Skip year-like numbers (>1000)
                        if harmonic > 1000:
                            continue
                        
                        if len(row) >= 9:
                            clean_row = [str(cell).strip() if cell is not None else "" for cell in row[:9]]
                            # Add page number as the last column
                            clean_row.append(page_number)
                            tables[active_table].append(clean_row)
                    except (ValueError, IndexError):
                        continue

def _extract_text_data(text, tables, active_table, page_number):
    """Helper function to extract text-based data with page tracking"""
    text_data = extract_table_data_from_text(text, page_number=page_number)
    if text_data:
        existing_harmonics = {int(row[0]) for row in tables[active_table] if row and str(row[0]).isdigit()}
        for new_row in text_data:
            try:
                harmonic_value = int(new_row[0])
                
                # ADDITIONAL FILTER: Ensure we only add valid harmonics (2-50)
                if 2 <= harmonic_value <= 50 and harmonic_value not in existing_harmonics:
                    tables[active_table].append(new_row)
            except (ValueError, IndexError):
                continue

def _check_boundary_hit(upper_text, active_table):
    """Helper function to check if section boundary is hit"""
    active_upper = active_table.upper()
    boundaries = SECTION_BOUNDARIES.get(active_upper, [])
    
    # For Harmonic Current Daily, exclude "HARMONIC 5:" from boundaries
    if active_upper == "HARMONIC CURRENT DAILY":
        boundaries = [b for b in boundaries if "HARMONIC 5" not in b.upper()]
    
    return any(boundary in upper_text for boundary in boundaries)

def extract_tables_from_pdf(file):
    """Extract all harmonic tables from PDF starting from page 2 with comprehensive page tracking"""
    tables = {table_name: [] for table_name in SUPPORTED_TABLES}
    
    try:
        with pdfplumber.open(file if isinstance(file, str) else file) as pdf:
            active_table = None
            
            # Skip first page as requested
            for page_num, page in enumerate(pdf.pages):
                if page_num == 0:  # Skip first page
                    continue
                    
                page_text = page.extract_text() or ""
                upper_text = page_text.upper()
                page_tables = page.extract_tables()
                
                # Actual page number (1-indexed)
                actual_page_num = page_num + 1
                
                logger.info(f"Processing page {actual_page_num} for table extraction")
                
                # Check for table headers
                for table_name in tables:
                    table_name_upper = table_name.upper()
                    if table_name_upper in upper_text:
                        start_idx = upper_text.find(table_name_upper)
                        end_idx = len(page_text)
                        
                        # Find section boundaries
                        for boundary in SECTION_BOUNDARIES.get(table_name_upper, []):
                            boundary_idx = upper_text.find(boundary, start_idx + len(table_name))
                            if boundary_idx != -1:
                                end_idx = min(end_idx, boundary_idx)
                        
                        section_text = page_text[start_idx:end_idx]
                        active_table = table_name
                        
                        logger.info(f"Found table '{table_name}' on page {actual_page_num}")
                        
                        # Extract structured tables with page tracking
                        initial_count = len(tables[active_table])
                        _extract_structured_data(page_tables, tables, active_table, actual_page_num)
                        new_count = len(tables[active_table])
                        
                        logger.info(f"Extracted {new_count - initial_count} structured rows from page {actual_page_num}")
                        
                        # Extract from text as fallback with page tracking
                        initial_count = len(tables[active_table])
                        _extract_text_data(section_text, tables, active_table, actual_page_num)
                        new_count = len(tables[active_table])
                        
                        logger.info(f"Extracted {new_count - initial_count} text rows from page {actual_page_num}")
                        continue
                
                # Continue extracting for active table
                if active_table and not _check_boundary_hit(upper_text, active_table):
                    initial_count = len(tables[active_table])
                    _extract_structured_data(page_tables, tables, active_table, actual_page_num)
                    _extract_text_data(page_text, tables, active_table, actual_page_num)
                    new_count = len(tables[active_table])
                    
                    if new_count > initial_count:
                        logger.info(f"Continued extracting {new_count - initial_count} rows for '{active_table}' from page {actual_page_num}")
                else:
                    # Only reset active_table if we actually hit a real boundary
                    if active_table and _check_boundary_hit(upper_text, active_table):
                        if active_table == "Harmonic Current Daily" and "HARMONIC 5:" in upper_text:
                            # Continue processing this page for the current table
                            initial_count = len(tables[active_table])
                            _extract_structured_data(page_tables, tables, active_table, actual_page_num)
                            _extract_text_data(page_text, tables, active_table, actual_page_num)
                            new_count = len(tables[active_table])
                            
                            if new_count > initial_count:
                                logger.info(f"Special case: Continued extracting {new_count - initial_count} rows for '{active_table}' from page {actual_page_num}")
                        else:
                            logger.info(f"Boundary hit for '{active_table}' on page {actual_page_num}, resetting active table")
                            active_table = None

    except Exception as e:
        logger.error(f"Error processing PDF: {str(e)}")
    
    # Log final extraction results
    for table_name, table_data in tables.items():
        if table_data:
            page_numbers = [row[-1] for row in table_data if len(row) > 9]
            unique_pages = set(page_numbers)
            logger.info(f"Table '{table_name}': {len(table_data)} rows extracted from pages: {sorted(unique_pages)}")
        else:
            logger.warning(f"Table '{table_name}': No data extracted")
    
    return tables

def process_table_data(table_data, table_name=None):
    """Process and validate table data with page number support - preserve all values"""
    columns = CURRENT_COLUMNS if table_name and "Current" in table_name else VOLTAGE_COLUMNS
    
    if not table_data:
        return pd.DataFrame(columns=columns)

    try:
        df = pd.DataFrame(table_data, columns=columns)
        
        # Only convert Harmonic column to numeric for filtering
        df['Harmonic'] = pd.to_numeric(df['Harmonic'], errors='coerce')
        df = df.dropna(subset=['Harmonic'])
        
        # CRITICAL FILTER: Remove fundamental frequency and invalid harmonics
        df = df[df['Harmonic'] != 1]
        df = df[(df['Harmonic'] >= 2) & (df['Harmonic'] <= 50)]
        
        df = df.drop_duplicates(subset=['Harmonic', 'Time Percent Limit[%]'])
        
        found_harmonics = set(df['Harmonic'].astype(int))
        missing = sorted(EXPECTED_HARMONICS - found_harmonics)
        
        if missing:
            logger.info(f"Missing harmonics in {table_name}: {missing[:10]}{'...' if len(missing) > 10 else ''}")
        
        if 'Page_Number' in df.columns:
            page_distribution = df['Page_Number'].value_counts().to_dict()
            logger.info(f"Page distribution for {table_name}: {page_distribution}")
        
        # Convert only essential numeric columns, preserve measured values as-is
        essential_cols = ["Harmonic", "Time Percent Limit[%]", "Reg Max[%]"]
        for col in essential_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        
        # Keep measured values as original (including zeros and strings)
        return df.dropna(subset=essential_cols)
    
    except Exception as e:
        logger.error(f"Error processing table data: {str(e)}")
        return pd.DataFrame(columns=columns)

def split_table(df):
    """Split table by time limits and odd/even harmonics"""
    if df.empty:
        return {"95": (pd.DataFrame(), pd.DataFrame()), "99": (pd.DataFrame(), pd.DataFrame())}
    
    def split_odd_even(df_subset):
        if df_subset.empty:
            return pd.DataFrame(), pd.DataFrame()
        df_subset = df_subset.sort_values("Harmonic")
        odd = df_subset[df_subset["Harmonic"] % 2 == 1].reset_index(drop=True)
        even = df_subset[df_subset["Harmonic"] % 2 == 0].reset_index(drop=True)
        return odd, even
    
    df_95 = df[df["Time Percent Limit[%]"] == 95.0].copy()
    df_99 = df[df["Time Percent Limit[%]"] == 99.0].copy()
    
    return {"95": split_odd_even(df_95), "99": split_odd_even(df_99)}

def analyze_failures(df, table_name=None):
    """Identify and summarize all harmonic violations with comprehensive page tracking"""
    if df.empty:
        return pd.DataFrame()
    
    violations = []
    measured_cols = [col for col in df.columns if col.startswith('Measured_')]
    
    for _, row in df.iterrows():
        try:
            threshold = float(row['Reg Max[%]'])
            harmonic = int(row['Harmonic'])
            time_limit = row['Time Percent Limit[%]']
            
            # Get page number with fallback
            page_number = row.get('Page_Number', 'Unknown')
            if pd.isna(page_number) or page_number is None:
                page_number = 'Unknown'
            else:
                try:
                    page_number = int(page_number)
                except (ValueError, TypeError):
                    page_number = 'Unknown'
            
            for col in measured_cols:
                phase = col.split('_')[-1]  # Gets V1N/V2N/V3N or I1/I2/I3
                value = float(row[col])
                
                if value > threshold:
                    violations.append({
                        'Harmonic': harmonic,
                        'Phase': phase,
                        'Time Limit (%)': time_limit,
                        'Allowed (%)': threshold,
                        'Measured (%)': value,
                        'Exceedance (%)': round(value - threshold, 2),
                        'Page': page_number,
                        'Table': table_name if table_name else 'Unknown'
                    })
                    logger.info(f"Violation found: Harmonic {harmonic}, Phase {phase}, Page {page_number}, Table {table_name}")
        except (ValueError, KeyError) as e:
            logger.warning(f"Error processing row for violations: {str(e)}")
            continue
    
    violations_df = pd.DataFrame(violations)
    
    if not violations_df.empty:
        # Log violations summary with page information
        total_violations = len(violations_df)
        unique_pages = violations_df['Page'].nunique()
        page_distribution = violations_df['Page'].value_counts().to_dict()
        
        logger.info(f"Violations analysis for {table_name}: {total_violations} violations across {unique_pages} pages")
        logger.info(f"Page distribution: {page_distribution}")
    
    return violations_df

def highlight_fails_in_excel(df, ws, start_row=1):
    """Apply conditional formatting to highlight failed harmonics in Excel with page info"""
    if df.empty:
        return
    
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fail_font = Font(color='9C0006', bold=True)
    harmonic_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    reg_max_col = None
    measured_cols = []
    result_cols = []
    page_col = None
    
    for idx, col in enumerate(df.columns, 1):
        if "Reg Max" in col:
            reg_max_col = idx
        elif col.startswith(('Measured_')):
            measured_cols.append(idx)
        elif col.startswith('Result_'):
            result_cols.append(idx)
        elif col == "Page_Number":
            page_col = idx
    
    if not reg_max_col or not measured_cols:
        return
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        if r_idx == start_row:
            continue
            
        try:
            threshold = float(ws.cell(row=r_idx, column=reg_max_col).value)
        except (ValueError, TypeError):
            threshold = 0.0
        
        for i, m_col in enumerate(measured_cols):
            try:
                cell = ws.cell(row=r_idx, column=m_col)
                value = float(cell.value) if cell.value else 0.0
                
                result_value = None
                if i < len(result_cols):
                    result_cell = ws.cell(row=r_idx, column=result_cols[i])
                    result_value = str(result_cell.value).lower() if result_cell.value else ""
                
                if value > threshold or (result_value and "fail" in result_value):
                    cell.fill = fail_fill
                    cell.font = fail_font
                    
                    harmonic_cell = ws.cell(row=r_idx, column=1)
                    harmonic_cell.fill = harmonic_fill
                    
                    # Highlight page number for violations
                    if page_col:
                        page_cell = ws.cell(row=r_idx, column=page_col)
                        page_cell.fill = fail_fill
                        page_cell.font = fail_font
            except:
                continue

def parse_filename_for_sheet_name(filename):
    """Parse filename to extract day number and time period for concise sheet naming"""
    filename_upper = filename.upper()
    
    if "7" in filename_upper and "DAY" in filename_upper:
        return "7Days"
    
    day_pattern = re.search(r'DAY\s*(\d+)\s*(DAY|NIGHT)', filename_upper)
    if day_pattern:
        day_num = day_pattern.group(1)
        period = "D" if "DAY" in day_pattern.group(2) else "N"
        return f"{day_num}{period}"
    
    day_only_pattern = re.search(r'DAY\s*(\d+)', filename_upper)
    if day_only_pattern:
        return f"{day_only_pattern.group(1)}D"
    
    clean_name = re.sub(r'[^\w]', '', filename.replace('.pdf', ''))
    return clean_name[:4]

def get_table_abbreviation(table_name):
    """Convert full table names to abbreviations for concise sheet naming"""
    table_upper = table_name.upper()
    
    if "CURRENT" in table_upper:
        prefix = "I"
    elif "VOLTAGE" in table_upper:
        prefix = "V"
    else:
        prefix = "X"
    
    if "FULL TIME RANGE" in table_upper:
        suffix = "F"
    elif "DAILY" in table_upper:
        suffix = "D"
    else:
        suffix = "X"
    
    return f"{prefix}{suffix}"

def create_excel_download(tables_data, filename):
    """Create Excel file with all tables with highlighting and split by odd/even harmonics"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for table_name, table_data in tables_data.items():
            if table_data:
                df = process_table_data(table_data, table_name)
                if not df.empty:
                    split_dfs = split_table(df)
                    
                    table_prefix = "I" if "Current" in table_name else "V"
                    table_suffix = "D" if "Daily" in table_name else "F"
                    
                    for limit in ["95", "99"]:
                        odd_df, even_df = split_dfs[limit]
                        
                        for df_data, suffix in [(odd_df, 'O'), (even_df, 'E')]:
                            if not df_data.empty:
                                sheet_name = f"H_{table_prefix}{table_suffix}_{limit}_{suffix}"[:31]
                                df_data.to_excel(writer, sheet_name=sheet_name, index=False)
                                
                                workbook = writer.book
                                worksheet = workbook[sheet_name]
                                highlight_fails_in_excel(df_data, worksheet, start_row=2)
    
    output.seek(0)
    return output.getvalue()

def create_bulk_excel_download(all_files_data):
    """Create Excel file with all PDFs using concise sheet naming and highlighting"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_file_map = {}
        
        for file_name, tables_data in all_files_data.items():
            file_prefix = parse_filename_for_sheet_name(file_name)
            
            for table_name, table_data in tables_data.items():
                if table_data:
                    df = process_table_data(table_data, table_name)
                    if not df.empty:
                        table_abbrev = get_table_abbreviation(table_name)
                        sheet_name = f"{file_prefix}_H_{table_abbrev}"
                        
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        
                        original_sheet_name = sheet_name
                        counter = 1
                        while sheet_name in sheet_file_map:
                            sheet_name = f"{original_sheet_name}_{counter}"
                            if len(sheet_name) > 31:
                                truncated = original_sheet_name[:31-len(f"_{counter}")]
                                sheet_name = f"{truncated}_{counter}"
                            counter += 1
                        
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheet_file_map[sheet_name] = file_name
                        
                        workbook = writer.book
                        worksheet = workbook[sheet_name]
                        highlight_fails_in_excel(df, worksheet, start_row=1)
    
    output.seek(0)
    wb = load_workbook(output)
    
    for sheet_name, file_name in sheet_file_map.items():
        ws = wb[sheet_name]
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=f"File: {file_name}")
    
    output2 = BytesIO()
    wb.save(output2)
    output2.seek(0)
    return output2.getvalue()

def create_bulk_word_download(all_files_data):
    """Create Word document with all PDFs using the same logic as Excel export"""
    from datetime import datetime
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # Start HTML document
    html_content = f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>Bulk Harmonic Analysis Report</title>
        <style>
            @page {{
                size: A4 landscape;
                margin: 0.5in 0.3in;
            }}
            
            body {{ 
                font-family: Arial, sans-serif; 
                font-size: 9pt; 
                margin: 0;
                line-height: 1.1;
            }}
            
            .page-container {{
                page-break-after: always;
                min-height: 95vh;
                max-height: 95vh;
                overflow: hidden;
                padding: 10px 0;
            }}
            
            .page-container:last-child {{
                page-break-after: avoid;
            }}
            
            .cover-page {{
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
                height: 95vh;
                text-align: center;
                background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
                border: 3px solid #0d6efd;
                border-radius: 15px;
                margin: 20px;
                padding: 40px;
            }}
            
            .file-title {{
                text-align: center;
                font-weight: bold;
                font-size: 16pt;
                color: #1e3a8a;
                margin: 20px 0;
                padding: 15px;
                border: 2px solid #3b82f6;
                background-color: #f1f5f9;
                border-radius: 8px;
            }}
            
            .table-title {{
                text-align: center;
                font-weight: bold;
                text-decoration: underline;
                margin: 15px 0 10px 0;
                font-size: 12pt;
                color: #1e40af;
            }}
            
            table {{ 
                border-collapse: collapse; 
                width: 100%; 
                font-size: 8pt;
                margin: 10px 0;
                page-break-inside: avoid;
                max-height: 70vh;
            }}
            
            th, td {{ 
                border: 1px solid #000; 
                padding: 3px 4px; 
                text-align: center;
                vertical-align: middle;
                word-wrap: break-word;
                line-height: 1.0;
            }}
            
            th {{ 
                background-color: #dbeafe !important;
                font-weight: bold; 
                font-size: 8pt;
                color: #1e40af;
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }}
            
            /* Even harmonics styling */
            .even-harmonic, .even-limit, .even-r, .even-y, .even-b {{
                background-color: #fff2cc !important;
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }}
            
            /* Odd harmonics styling */
            .odd-harmonic, .odd-limit, .odd-r, .odd-y, .odd-b {{
                background-color: #e1f5fe !important;
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }}
            
            /* Current table variations */
            .current-table .even-harmonic, .current-table .even-limit, 
            .current-table .even-r, .current-table .even-y, .current-table .even-b {{
                background-color: #e0f2fe !important;
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }}
            
            .current-table .odd-harmonic, .current-table .odd-limit,
            .current-table .odd-r, .current-table .odd-y, .current-table .odd-b {{
                background-color: #dcfce7 !important;
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }}
            
            .remarks {{
                background-color: #f9fafb !important;
                color: #374151;
                font-size: 7pt;
                text-align: left;
                padding-left: 5px;
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }}

            /* Violation highlighting */
            .violation {{ 
                background-color: #fecaca !important; 
                color: #dc2626 !important; 
                font-weight: bold; 
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }}

            .within-limits {{
            color: #16a34a !important;  /* Green color */
            font-weight: bold;
            }}

            .exceeds-limits {{
            color: #dc2626 !important;  /* Red color */
            font-weight: bold;
            }}
            
            .summary-box {{
                background-color: #e3f2fd;
                border: 1px solid #1976d2;
                border-radius: 5px;
                padding: 15px;
                margin: 15px 0;
                text-align: center;
            }}
        </style>
    </head>
    <body>
        <!-- ENHANCED COVER PAGE -->
        <div class="page-container cover-page">
            <h1 style="font-size: 24pt; color: #1e3a8a; margin-bottom: 10px;">
                📊 BULK HARMONIC ANALYSIS REPORT
            </h1>
            <div class="summary-box">
                <h2 style="color: #1565c0; margin-bottom: 20px;">Report Summary</h2>
                <p style="font-size: 14pt; margin: 10px 0;"><strong>📅 Generated Date:</strong> {current_date}</p>
                <p style="font-size: 14pt; margin: 10px 0;"><strong>📁 Total Files Analyzed:</strong> {len(all_files_data)}</p>
                <p style="font-size: 14pt; margin: 10px 0;"><strong>🔍 Analysis Type:</strong> IEEE 519-2022 Harmonic Compliance</p>
            </div>
            <div style="margin-top: 40px; font-size: 12pt; color: #666;">
                <p>This report contains individual harmonic distortion analysis for:</p>
                <ul style="text-align: left; display: inline-block; margin-top: 15px;">
                    <li>Voltage Harmonic Distortion (3sec & 10min)</li>
                    <li>Current Harmonic Distortion (3sec & 10min)</li>
                    <li>IEEE 519-2022 Compliance Assessment</li>
                    <li>Violation Detection & Analysis</li>
                </ul>
            </div>
        </div>
    """
    
    # Process each file using the SAME LOGIC as Excel export
    for file_index, (file_name, tables_data) in enumerate(all_files_data.items()):
        # Add file separator
        html_content += f"""
        <div class="page-container">
            <div class="file-title">
                File {file_index + 1}: {file_name}
            </div>
        """
        
        # Process each table using EXACT Excel logic
        for table_name, table_data in tables_data.items():
            if table_data:
                # Use the SAME processing as Excel export
                df = process_table_data(table_data, table_name)
                if not df.empty:
                    split_dfs = split_table(df)
                    
                    # Get table display info
                    table_prefix = "Current" if "Current" in table_name else "Voltage"
                    table_suffix = "Daily" if "Daily" in table_name else "Full Time Range"
                    
                    for limit in ["95", "99"]:
                        if limit in split_dfs:
                            odd_df, even_df = split_dfs[limit]
                            
                            if not odd_df.empty or not even_df.empty:
                                # Add table to SAME page as file title if it fits
                                display_title = f"Individual {table_prefix} Harmonic distortion ({table_suffix}) {limit}th percentile"
                                
                                html_content += f"""
                                <div class="table-title">{display_title}</div>
                                """
                                
                                # Generate table using Excel data structure
                                table_html = _generate_excel_compatible_table_html(
                                    odd_df, even_df, table_name, limit
                                )
                                html_content += table_html
        
        html_content += "</div>"  # Close file page
    
    html_content += "</body></html>"
    
    return html_content.encode('utf-8')

def is_violation(value, limit):
    try:
        return value != "" and limit != "" and float(value) > float(limit)
    except (ValueError, TypeError):
        return False
    

def create_enhanced_excel_download(tables_data, filename):
    """Create Excel file with both individual harmonic tables and THD/TDD summary tables"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Add individual harmonic tables
        for table_name, table_data in tables_data.items():
            if table_data:
                df = process_table_data(table_data, table_name)
                if not df.empty:
                    split_dfs = split_table(df)
                    
                    table_prefix = "I" if "Current" in table_name else "V"
                    table_suffix = "D" if "Daily" in table_name else "F"
                    
                    for limit in ["95", "99"]:
                        odd_df, even_df = split_dfs[limit]
                        
                        for df_data, suffix in [(odd_df, 'O'), (even_df, 'E')]:
                            if not df_data.empty:
                                sheet_name = f"H_{table_prefix}{table_suffix}_{limit}_{suffix}"[:31]
                                df_data.to_excel(writer, sheet_name=sheet_name, index=False)
                                
                                workbook = writer.book
                                worksheet = workbook[sheet_name]
                                highlight_fails_in_excel(df_data, worksheet, start_row=2)
    
    output.seek(0)
    return output.getvalue()


def _generate_excel_compatible_table_html(odd_df, even_df, table_name, limit):
    """Generate HTML table using the EXACT same logic as Excel export - show all values including zeros"""
    
    table_class = "current-table" if "Current" in table_name else "voltage-table"
    
    html = f'<table class="{table_class}">'
    
    html += """
    <thead>
        <tr>
            <th rowspan="2" class="even-harmonic">Even Harmonic</th>
            <th rowspan="2" class="even-limit">Recommended Limits (%)</th>
            <th colspan="3">Measured max [%]</th>
            <th rowspan="2" class="odd-harmonic">Odd Harmonic</th>
            <th rowspan="2" class="odd-limit">Recommended Limits (%)</th>
            <th colspan="3">Measured max [%]</th>
            <th rowspan="2" class="remarks">Remarks</th>
        </tr>
        <tr>
            <th class="even-r">R</th>
            <th class="even-y">Y</th>
            <th class="even-b">B</th>
            <th class="odd-r">R</th>
            <th class="odd-y">Y</th>
            <th class="odd-b">B</th>
        </tr>
    </thead>
    <tbody>
    """
    
    even_data = {}
    odd_data = {}
    
    if not even_df.empty:
        for _, row in even_df.iterrows():
            harmonic = int(row['Harmonic'])
            measured_cols = ['Measured_I1', 'Measured_I2', 'Measured_I3'] if 'Current' in table_name else ['Measured_V1N', 'Measured_V2N', 'Measured_V3N']
            
            even_data[harmonic] = {
                'limit': row['Reg Max[%]'] if pd.notna(row['Reg Max[%]']) else "",
                'values': [
                    row[measured_cols[0]] if pd.notna(row[measured_cols[0]]) else "",
                    row[measured_cols[1]] if pd.notna(row[measured_cols[1]]) else "",
                    row[measured_cols[2]] if pd.notna(row[measured_cols[2]]) else ""
                ]
            }
    
    if not odd_df.empty:
        for _, row in odd_df.iterrows():
            harmonic = int(row['Harmonic'])
            measured_cols = ['Measured_I1', 'Measured_I2', 'Measured_I3'] if 'Current' in table_name else ['Measured_V1N', 'Measured_V2N', 'Measured_V3N']
            
            odd_data[harmonic] = {
                'limit': row['Reg Max[%]'] if pd.notna(row['Reg Max[%]']) else "",
                'values': [
                    row[measured_cols[0]] if pd.notna(row[measured_cols[0]]) else "",
                    row[measured_cols[1]] if pd.notna(row[measured_cols[1]]) else "",
                    row[measured_cols[2]] if pd.notna(row[measured_cols[2]]) else ""
                ]
            }
    
    even_harmonics = list(range(2, 51, 2))
    odd_harmonics = list(range(3, 50, 2))
    max_rows = max(len(even_harmonics), len(odd_harmonics))
    
    has_violations = False
    
    for i in range(max_rows):
        html += "<tr>"
        
        # Even harmonic column
        if i < len(even_harmonics):
            even_harmonic = even_harmonics[i]
            even_info = even_data.get(even_harmonic, {'limit': "", 'values': ["", "", ""]})
            
            html += f'<td class="even-harmonic">{even_harmonic}</td>'
            html += f'<td class="even-limit"><strong>{even_info["limit"]}</strong></td>'
            
            for j, value in enumerate(even_info['values']):
                cell_class = f"even-{'ryb'[j]}"
                
                # Check for violations - only if both value and limit exist and are numeric
                try:
                    if (value != "" and even_info['limit'] != "" and 
                        float(value) > float(even_info['limit'])):
                        cell_class += " violation"
                        has_violations = True
                except (ValueError, TypeError):
                    pass
                
                # Display ALL values including zeros and empty strings
                display_value = str(value) if value != "" else ""
                html += f'<td class="{cell_class}">{display_value}</td>'
        else:
            html += '<td class="even-harmonic"></td><td class="even-limit"></td>'
            html += '<td class="even-r"></td><td class="even-y"></td><td class="even-b"></td>'
        
        # Odd harmonic column
        if i < len(odd_harmonics):
            odd_harmonic = odd_harmonics[i]
            odd_info = odd_data.get(odd_harmonic, {'limit': "", 'values': ["", "", ""]})
            
            html += f'<td class="odd-harmonic">{odd_harmonic}</td>'
            html += f'<td class="odd-limit"><strong>{odd_info["limit"]}</strong></td>'
            
            for j, value in enumerate(odd_info['values']):
                cell_class = f"odd-{'ryb'[j]}"
                cell_style = ""
                try:
                    if (value != "" and odd_info['limit'] != "" and float(value) > float(odd_info['limit'])):
                        cell_style = ' style="color:#dc2626; background-color:#fecaca; font-weight:bold;"'
                        has_violations = True
                except (ValueError, TypeError):
                    pass
                display_value = str(value) if value != "" else ""
                html += f'<td class="{cell_class}"{cell_style}>{display_value}</td>'
        else:
            html += '<td class="odd-harmonic"></td><td class="odd-limit"></td>'
            html += '<td class="odd-r"></td><td class="odd-y"></td><td class="odd-b"></td>'
        
        # Remarks column
        if i == 0:
            # Check for any violations in this table
            has_violations = False
            for harmonic_info in list(even_data.values()) + list(odd_data.values()):
                if harmonic_info['limit']:
                    for value in harmonic_info['values']:
                        if is_violation(value, harmonic_info['limit']):
                            has_violations = True
                            break
                if has_violations:
                    break
            
            if has_violations:
                html += '<td class="remarks exceeds-limits" style="color:#dc2626">Values exceeding limits detected</td>'
            else:
                html += '<td class="remarks within-limits" style="color:#008000">All values within limits</td>'
        else:
            html += '<td class="remarks"></td>'
        html += "</tr>"
    
    html += "</tbody></table>"
    return html